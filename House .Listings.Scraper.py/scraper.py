import requests
import pandas as pd
import json
import time


API_URL = "https://api2.realtor.ca/Listing.svc/PropertySearch_Post"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Referer": "https://www.realtor.ca/",
    "Content-Type": "application/x-www-form-urlencoded",
}


PAYLOAD = {
    "ZoomLevel": "12",
    "LatitudeMax": "49.2800",
    "LongitudeMax": "-122.6200",
    "LatitudeMin": "49.2200",
    "LongitudeMin": "-122.7500",
    "Sort": "6-D",
    "PropertyTypeGroupID": "1",
    "BedRange": "3-0",
    "BathRange": "2-0",
    "TransactionTypeId": "2",
    "PropertySearchTypeId": "1",
    "Currency": "CAD",
    "RecordsPerPage": "50",
    "ApplicationId": "1",
    "CultureId": "1",
    "Version": "7.0",
    "CurrentPage": "1",
}

def fetch_listings():
    print("Fetching listings from Realtor.ca...")
    response = requests.post(API_URL, headers=HEADERS, data=PAYLOAD, timeout=15)
    response.raise_for_status()
    return response.json()

def parse_listings(data):
    results = data.get("Results", [])
    listings = []

    for item in results:
        prop = item.get("Property", {})
        address = item.get("PostalCode", "")
        address_full = prop.get("Address", {})

        beds = prop.get("BedroomTotal", "N/A")
        baths = prop.get("BathroomTotal", "N/A")
        price = prop.get("Price", "N/A")
        listing_id = item.get("MlsNumber", "")
        url = f"https://www.realtor.ca/{item.get('RelativeURLEn', '')}"
        prop_type = prop.get("Type", "N/A")
        sqft = prop.get("SizeInterior", "N/A")

        city = address_full.get("City", "")
        street = address_full.get("AddressText", "N/A")


        if "Pitt Meadows" not in city and "Pitt Meadows" not in street:
            continue

        listings.append({
            "MLS #": listing_id,
            "Address": street,
            "City": city,
            "Price": price,
            "Bedrooms": beds,
            "Bathrooms": baths,
            "Type": prop_type,
            "Size (sqft)": sqft,
            "URL": url,
        })

    return listings

def save_to_excel(listings, filename="pitt_meadows_listings.xlsx"):
    if not listings:
        print("No listings found matching the criteria.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pitt Meadows Listings"


    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Pitt Meadows — Residential Listings  |  3 Bed / 2 Bath+"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"Scraped from Realtor.ca  •  {len(listings)} listings found"
    sub.font = Font(name="Arial", size=10, italic=True, color="FFFFFF")
    sub.fill = PatternFill("solid", fgColor="2E75B6")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # --- Column headers ---
    headers = ["MLS #", "Address", "City", "Price", "Bedrooms", "Bathrooms", "Type", "Size (sqft)", "Listing URL"]
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 22


    for row_num, listing in enumerate(listings, 4):
        row_data = [
            listing["MLS #"],
            listing["Address"],
            listing["City"],
            listing["Price"],
            listing["Bedrooms"],
            listing["Bathrooms"],
            listing["Type"],
            listing["Size (sqft)"],
            listing["URL"],
        ]
        fill_color = "F2F7FC" if row_num % 2 == 0 else "FFFFFF"
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # Right-align price
            if col_num == 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            # Center beds/baths/type
            if col_num in (5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 18

    # --- Column widths ---
    col_widths = [14, 38, 22, 14, 11, 12, 16, 13, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


    ws.freeze_panes = "A4"

    wb.save(filename)
    print(f"Saved {len(listings)} listings to '{filename}'")

def main():
    try:
        data = fetch_listings()
        listings = parse_listings(data)
        print(f"Found {len(listings)} listings in Pitt Meadows (3 bed / 2 bath+)")
        save_to_excel(listings)
    except requests.exceptions.RequestException as e:
        print(f"Request failed: {e}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()